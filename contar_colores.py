import openpyxl
from PIL import ImageColor

# Cargar el archivo Excel original
archivo = 'C:\\Users\\Sindy Mejia\\Documents\\Reporte 25+ Color\\Reporte25-CountColor\\Reporte 25+ Contar.xlsx'
wb = openpyxl.load_workbook(archivo)
hoja = wb.active

# Colores a buscar
color_rojo = ImageColor.getrgb("#FF0000")  # Rojo
color_naranja = ImageColor.getrgb("#FF9900")  # Naranja

# Diccionario para contar celdas por columna
conteo_columnas = {}

# Obtener los encabezados
encabezados = [celda.value for celda in hoja[1]]

# Inicializar el conteo por cada encabezado
for header in encabezados:
    conteo_columnas[header] = {'rojo': 0, 'naranja': 0}

# Recorrer las celdas (empezando desde la segunda fila)
for fila in hoja.iter_rows(min_row=2):  # Saltar la fila de encabezados
    for celda in fila:
        if celda.fill and celda.fill.fgColor:
            color_actual = celda.fill.fgColor
            if color_actual:
                if color_actual.type == 'rgb':
                    hex_color = color_actual.rgb
                    if hex_color == '00000000':
                        continue
                elif color_actual.type == 'theme':
                    continue
                elif color_actual.type == 'indexed':
                    continue

                if len(hex_color) == 8:  # RGB en formato ARGB
                    hex_color = hex_color[2:]  # Ignorar el componente alfa
                elif len(hex_color) != 6:
                    continue

                # Obtener el encabezado correspondiente
                col_index = celda.column - 1
                header = encabezados[col_index]

                # Comparar los colores y contar
                if hex_color[:6].upper() == f'{color_rojo[0]:02X}{color_rojo[1]:02X}{color_rojo[2]:02X}':
                    conteo_columnas[header]['rojo'] += 1
                elif hex_color[:6].upper() == f'{color_naranja[0]:02X}{color_naranja[1]:02X}{color_naranja[2]:02X}':
                    conteo_columnas[header]['naranja'] += 1

# Crear un nuevo archivo Excel para los resultados
nuevo_archivo = 'C:\\Users\\Sindy Mejia\\Documents\\Reporte 25+ Color\\Reporte25-CountColor\\resultado_colores.xlsx'
wb_nuevo = openpyxl.Workbook()
hoja_nueva = wb_nuevo.active
hoja_nueva.title = "Resultados"

# Escribir encabezados en el nuevo archivo
hoja_nueva.append(['Encabezado', 'Celdas Rojas', 'Celdas Naranjas'])

# Escribir los conteos por cada encabezado
for header, conteo in conteo_columnas.items():
    hoja_nueva.append([header, conteo['rojo'], conteo['naranja']])

# Calcular totales
total_rojo = sum(conteo['rojo'] for conteo in conteo_columnas.values())
total_naranja = sum(conteo['naranja'] for conteo in conteo_columnas.values())

# Agregar totales al final
hoja_nueva.append(['Total', total_rojo, total_naranja])

# Guardar el nuevo archivo
wb_nuevo.save(nuevo_archivo)
print(f'Archivo generado: {nuevo_archivo}')
